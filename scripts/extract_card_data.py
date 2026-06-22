"""Extract the Revised-Edition card catalog (base + 5 expansions) from the Agricola Database.

Source of truth: ``Rulebooks/Agricola Database.xlsx`` (sheet ``Database``), filtered to
``Edition == "Revised"`` and ``Base/Expansion`` in the in-scope set below. Scope:

    Base (Revised), Artifex, Bubulcus, Corbarius, Dulcinaria, Consul Dirigens

Writes two flat JSON catalogs under ``agricola/cards/data/``:

  * ``revised_occupations.json``        — 336 cards (pure occupations)
  * ``revised_minor_improvements.json`` — 336 cards (332 minor + 4 dual cards, see below)

The database tags four cards with compound ``Type`` labels ("Minor/Major Improvement" ×3:
Oriental Fireplace, Earth Oven, Large Pottery; "Major Improvement/Occupation" ×1: Witches' Dance
Floor). All four are acquired via the *Minor Improvement* action (the compendium is explicit for
Witches' Dance Floor, D025: "You can play it only via a 'minor Improvement' action"), so we treat
them as plain minor improvements — their dual nature is already spelled out in each card's text.
Their stored ``type`` is normalized to "Minor Improvement".

The two files carry different columns on purpose: across this whole scope occupations never have
cost/VP/prerequisite/passing-left, so the occupation file omits those; minor improvements have no
player-count restriction (all legal in a 2-player game), so the minor file omits ``players``.

``(expansion, deck, number)`` is the unique key — deck letters and card numbers repeat across
expansions (Artifex=A, Bubulcus=B, …), so ``expansion`` is part of the key, not decoration.

Deliberately EXCLUDED (neither an occupation nor a minor improvement): the 10 revised-base Major
Improvements and the 24 Consul Dirigens Parent cards (decks PR/PS). Add a separate file if wanted.

Re-run after editing the xlsx:

    ~/miniconda3/bin/python scripts/extract_card_data.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "Rulebooks" / "Agricola Database.xlsx"
OUT_DIR = REPO / "agricola" / "cards" / "data"

SCOPE = {
    "Base (Revised)",
    "Artifex Expansion",
    "Bubulcus Expansion",
    "Corbarius Expansion",
    "Dulcinaria Expansion",
    "Consul Dirigens Expansion",
}

# Implementation-tracking field (not from the database). New cards default to "todo"; existing
# status is preserved across re-extracts so editing the database (or the workbook) never resets
# progress. Keyed by (expansion, deck, number).
STATUS_VALUES = ("implemented", "todo", "wontfix")
DEFAULT_STATUS = "todo"

# The four compound-type cards are all played via the Minor Improvement action, so they fold into
# the minor-improvement file with their `type` normalized to "Minor Improvement" (see docstring).
OCC_TYPES = {"Occupation"}
MIN_TYPES = {"Minor Improvement", "Minor/Major Improvement", "Major Improvement/Occupation"}

# Columns kept per file (tailored — see module docstring). `status` is the tracking field.
OCC_COLS = ["expansion", "deck", "number", "type", "status",
            "players", "name", "card_category", "text"]
MIN_COLS = ["expansion", "deck", "number", "type", "status", "name", "cost", "vps",
            "prerequisites", "passing_left", "card_category", "text"]

HEADER_KEY = {
    "Base/Expansion": "expansion",
    "Deck": "deck",
    "Number": "number",
    "Type": "type",
    "Player(s)": "players",
    "Name": "name",
    "Cost": "cost",
    "VPs": "vps",
    "Prerequisites": "prerequisites",
    "Passing Left": "passing_left",
    "Card Category (Revised only)": "card_category",
    "Text": "text",
}


def _clean(value):
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _to_int(value):
    # Defensive: only collapse whole numbers; leave any string value (e.g. a future "1/3" VP) as-is.
    if isinstance(value, (int, float)):
        return int(value)
    return value


def load_scope_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Database"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in raw):
            continue
        if raw[idx["Edition"]] != "Revised":
            continue
        if raw[idx["Base/Expansion"]] not in SCOPE:
            continue
        record = {key: _clean(raw[idx[h]]) for h, key in HEADER_KEY.items()}
        record["number"] = _to_int(record["number"])
        record["vps"] = _to_int(record["vps"])
        rows.append(record)
    return rows


def load_existing_status(out_dir):
    """Map (expansion, deck, number) -> status from any catalogs already on disk."""
    status = {}
    for fname in ("revised_occupations.json", "revised_minor_improvements.json"):
        path = out_dir / fname
        if not path.exists():
            continue
        for card in json.loads(path.read_text(encoding="utf-8")):
            if "status" in card:
                status[(card["expansion"], card["deck"], card["number"])] = card["status"]
    return status


def write_catalog(rows, types, columns, out_path, normalize_type=None):
    cards = [r for r in rows if r["type"] in types]
    cards.sort(key=lambda r: (r["expansion"], r["deck"], r["number"]))
    if normalize_type is not None:
        cards = [{**card, "type": normalize_type} for card in cards]
    projected = [{col: card[col] for col in columns} for card in cards]
    out_path.write_text(json.dumps(projected, indent=2, ensure_ascii=False) + "\n",
                        encoding="utf-8")
    return len(projected)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = load_scope_rows()
    status_map = load_existing_status(OUT_DIR)
    for r in rows:
        r["status"] = status_map.get((r["expansion"], r["deck"], r["number"]), DEFAULT_STATUS)
    n_occ = write_catalog(rows, OCC_TYPES, OCC_COLS,
                          OUT_DIR / "revised_occupations.json")
    n_min = write_catalog(rows, MIN_TYPES, MIN_COLS,
                          OUT_DIR / "revised_minor_improvements.json",
                          normalize_type="Minor Improvement")
    excluded = [r for r in rows if r["type"] not in (OCC_TYPES | MIN_TYPES)]
    print(f"Wrote {n_occ} occupations and {n_min} minor improvements to {OUT_DIR}")
    print(f"Excluded {len(excluded)} non-occupation/non-minor rows "
          f"(majors + parent cards): " +
          ", ".join(sorted({r["type"] for r in excluded})))


if __name__ == "__main__":
    main()
