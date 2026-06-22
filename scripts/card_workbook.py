"""Build / sync the human-facing Excel workbook for the revised card catalog.

The JSON catalogs under ``agricola/cards/data/`` are canonical (the engine reads them). This
workbook is a convenience layer for browsing and for editing the ``status`` column:

    --build : JSON catalogs  ->  revised_cards.xlsx   (one sheet per type; filter dropdowns on
                                                       every column; a validated status dropdown;
                                                       status cells colour-coded)
    --sync  : revised_cards.xlsx  ->  JSON catalogs    (pulls ONLY the status column back; all
                                                        card data stays sourced from the database)

Typical loop: edit status in the workbook -> ``--sync`` -> the engine sees it. If you ever
re-run ``extract_card_data.py`` (e.g. the database changed), ``--sync`` FIRST so unsynced status
edits aren't lost, then ``--build`` to refresh the workbook.

    ~/miniconda3/bin/python scripts/card_workbook.py --build
    ~/miniconda3/bin/python scripts/card_workbook.py --sync
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from extract_card_data import OUT_DIR, STATUS_VALUES

XLSX = OUT_DIR / "revised_cards.xlsx"
# sheet title -> json filename
SHEETS = {
    "Occupations": OUT_DIR / "revised_occupations.json",
    "Minor Improvements": OUT_DIR / "revised_minor_improvements.json",
}
KEY_COLS = ("expansion", "deck", "number")

COL_WIDTHS = {
    "expansion": 22, "deck": 6, "number": 8, "type": 16, "status": 14, "players": 9,
    "name": 26, "cost": 22, "vps": 6, "prerequisites": 24, "passing_left": 12,
    "card_category": 26, "text": 95,
}
STATUS_FILLS = {  # live colour-coding via conditional formatting
    "implemented": PatternFill("solid", fgColor="C6EFCE"),  # green
    "todo":        PatternFill("solid", fgColor="FFEB9C"),  # amber
    "wontfix":     PatternFill("solid", fgColor="D9D9D9"),  # grey
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP_COLS = {"text", "name", "card_category", "prerequisites"}


def _load(path):
    return json.loads(path.read_text(encoding="utf-8"))


def build():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, path in SHEETS.items():
        cards = _load(path)
        columns = list(cards[0].keys())
        ws = wb.create_sheet(title)

        ws.append(columns)
        for col_idx, name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill, cell.font = HEADER_FILL, HEADER_FONT
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(name, 16)

        for card in cards:
            ws.append([card.get(c) for c in columns])

        for col_idx, name in enumerate(columns, start=1):
            if name in WRAP_COLS:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                        wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions  # filter dropdown on every column

        # validated status dropdown + live colour-coding on the status column
        status_letter = get_column_letter(columns.index("status") + 1)
        rng = f"{status_letter}2:{status_letter}{ws.max_row}"
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUS_VALUES),
                            allow_blank=False, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(rng)
        for value, fill in STATUS_FILLS.items():
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="equal", formula=['"%s"' % value], fill=fill))

    wb.save(XLSX)
    print(f"Built {XLSX} ({len(SHEETS)} sheets)")


def sync():
    total_changed = 0
    for title, path in SHEETS.items():
        wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
        ws = wb[title]
        rows = ws.iter_rows(values_only=True)
        header = list(next(rows))
        ix = {h: header.index(h) for h in (*KEY_COLS, "status")}
        wb_status = {}
        for r in rows:
            if r[ix["expansion"]] is None:
                continue
            key = (r[ix["expansion"]], r[ix["deck"]], int(r[ix["number"]]))
            value = r[ix["status"]]
            if value not in STATUS_VALUES:
                raise SystemExit(f"[{title}] invalid status {value!r} for {key}; "
                                 f"allowed: {STATUS_VALUES}")
            wb_status[key] = value

        cards = _load(path)
        changed = 0
        for card in cards:
            key = (card["expansion"], card["deck"], card["number"])
            if key in wb_status and wb_status[key] != card["status"]:
                card["status"] = wb_status[key]
                changed += 1
        path.write_text(json.dumps(cards, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        total_changed += changed
        print(f"[{title}] updated status on {changed} card(s)")
    print(f"Synced {total_changed} status change(s) into the JSON catalogs")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--build", action="store_true", help="JSON -> xlsx")
    g.add_argument("--sync", action="store_true", help="xlsx status column -> JSON")
    args = ap.parse_args()
    build() if args.build else sync()


if __name__ == "__main__":
    main()
